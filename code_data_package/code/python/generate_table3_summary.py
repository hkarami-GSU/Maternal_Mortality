from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
PACKAGE_ROOT = SCRIPT_DIR.parents[1]
OUTPUT_FILE = PACKAGE_ROOT / "outputs" / "tables" / "Table3_Global_Regional_Excess_Maternal_Mortality.xlsx"


@dataclass(frozen=True)
class RegionSpec:
    display: str
    file_key: str


REGIONS = [
    RegionSpec("African Region", "African Region"),
    RegionSpec("Region of the Americas", "Region of the Americas"),
    RegionSpec("Eastern Mediterranean Region", "Eastern Mediterranean"),
    RegionSpec("European Region", "European Region"),
    RegionSpec("South-East Asia Region", "Southeast Asia Region"),
    RegionSpec("Western Pacific Region", "Western Pacific Region"),
]


EXPECTED = {
    "Global": (195, 68489, 34706, 147118, 10154, 4568, 23744, 113, 133),
    "African Region": (47, 13789, 2920, 58232, 3547, 608, 13308, 15, 16),
    "Region of the Americas": (36, 9583, 7334, 13538, 2658, 1668, 3858, 23, 30),
    "Eastern Mediterranean Region": (22, 13886, 5216, 25751, 1130, 653, 2129, 19, 21),
    "European Region": (53, 1210, 606, 2022, 977, 724, 1321, 33, 40),
    "South-East Asia Region": (11, 28030, 17669, 43597, 643, 447, 1079, 7, 8),
    "Western Pacific Region": (26, 1991, 961, 3978, 1199, 468, 2049, 16, 18),
}


TRIPLET_RE = re.compile(
    r"^\s*(-?\d+(?:\.\d+)?)\s*\(\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\)\s*$"
)


def parse_triplet(value: object) -> tuple[float, float, float]:
    if value is None:
        raise ValueError("Missing triplet value")
    match = TRIPLET_RE.match(str(value))
    if not match:
        raise ValueError(f"Could not parse triplet value: {value!r}")
    return tuple(float(part) for part in match.groups())


def as_output_number(value: float) -> int | float:
    rounded = round(value)
    if abs(value - rounded) < 1e-9:
        return int(rounded)
    return value


def fmt_num(value: float) -> str:
    value = as_output_number(value)
    if isinstance(value, int):
        return f"{value:,}"
    return f"{value:,.1f}"


def fmt_triplet(median: float, lower: float, upper: float) -> str:
    return f"{fmt_num(median)} ({fmt_num(lower)}--{fmt_num(upper)})"


def read_formatted_workbook(path: Path) -> dict[str, tuple[float, float, float]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"{path} is empty")

    header = list(rows[0])
    country_col = header.index("State")
    total_col = header.index("Excess (LB,UB) Total")

    parsed: dict[str, tuple[float, float, float]] = {}
    for row in rows[1:]:
        if not row or row[country_col] in (None, ""):
            continue
        country = str(row[country_col]).strip()
        if country in parsed:
            raise ValueError(f"Duplicated country {country!r} in {path}")
        parsed[country] = parse_triplet(row[total_col])
    return parsed


def source_paths(region: RegionSpec) -> tuple[Path, Path]:
    death_path = (
        PACKAGE_ROOT
        / "outputs"
        / "maternal_deaths_workflow"
        / "excess_deaths_figures"
        / f"ExcessMaternalDeaths_{region.file_key}_Formatted.xlsx"
    )
    mmr_path = (
        PACKAGE_ROOT
        / "outputs"
        / "mmr_workflow"
        / "excess_deaths_figures"
        / f"ExcessDeaths_{region.file_key}_Formatted.xlsx"
    )
    return death_path, mmr_path


def add_triplets(values: list[tuple[float, float, float]]) -> tuple[float, float, float]:
    return tuple(sum(value[index] for value in values) for index in range(3))


def style_sheet(ws, freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def write_table3(ws, rows: list[dict[str, object]]) -> None:
    headers = [
        "Region",
        "Countries included",
        "Cumulative excess maternal deaths",
        "Deaths lower 95% UI",
        "Deaths upper 95% UI",
        "Formatted cumulative excess maternal deaths",
        "Aggregate excess MMR",
        "MMR lower 95% UI",
        "MMR upper 95% UI",
        "Formatted aggregate excess MMR",
        "Countries with detectable excess deaths",
        "Countries with detectable excess MMR",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[header] for header in headers])

    style_sheet(ws)
    set_widths(
        ws,
        {
            "A": 30,
            "B": 16,
            "C": 20,
            "D": 18,
            "E": 18,
            "F": 34,
            "G": 18,
            "H": 16,
            "I": 16,
            "J": 30,
            "K": 24,
            "L": 24,
        },
    )
    for col in ["B", "C", "D", "E", "G", "H", "I", "K", "L"]:
        for cell in ws[col][1:]:
            cell.number_format = "#,##0"


def write_country_level(ws, rows: list[dict[str, object]]) -> None:
    headers = [
        "Region",
        "Country",
        "Death median total",
        "Death lower total",
        "Death upper total",
        "MMR median total",
        "MMR lower total",
        "MMR upper total",
        "Detectable deaths",
        "Detectable MMR",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[header] for header in headers])
    style_sheet(ws)
    set_widths(
        ws,
        {
            "A": 30,
            "B": 34,
            "C": 18,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 18,
            "I": 18,
            "J": 18,
        },
    )
    for col in ["C", "D", "E", "F", "G", "H"]:
        for cell in ws[col][1:]:
            cell.number_format = "#,##0"


def write_checks(ws, checks: list[tuple[str, object]]) -> None:
    ws.append(["Check", "Value"])
    for check, value in checks:
        ws.append([check, value])
    style_sheet(ws)
    set_widths(ws, {"A": 48, "B": 100})


def main() -> None:
    country_rows: list[dict[str, object]] = []
    table_rows_by_region: list[dict[str, object]] = []
    checks: list[tuple[str, object]] = []
    all_countries: list[str] = []
    global_death_triplets: list[tuple[float, float, float]] = []
    global_mmr_triplets: list[tuple[float, float, float]] = []

    corrections = [
        "Eastern Mediterranean Region -> Eastern Mediterranean in source filenames",
        "South-East Asia Region -> Southeast Asia Region in source filenames",
    ]
    all_missing: list[str] = []
    all_duplicates: list[str] = []
    set_matches: list[bool] = []

    for region in REGIONS:
        death_path, mmr_path = source_paths(region)
        if not death_path.exists():
            raise FileNotFoundError(death_path)
        if not mmr_path.exists():
            raise FileNotFoundError(mmr_path)

        deaths = read_formatted_workbook(death_path)
        mmr = read_formatted_workbook(mmr_path)
        death_countries = set(deaths)
        mmr_countries = set(mmr)
        set_matches.append(death_countries == mmr_countries)

        missing_death = sorted(mmr_countries - death_countries)
        missing_mmr = sorted(death_countries - mmr_countries)
        if missing_death:
            all_missing.append(f"{region.display}: missing deaths for {', '.join(missing_death)}")
        if missing_mmr:
            all_missing.append(f"{region.display}: missing MMR for {', '.join(missing_mmr)}")

        combined_countries = sorted(death_countries | mmr_countries)
        seen: set[str] = set()
        for country in combined_countries:
            if country in seen:
                all_duplicates.append(f"{region.display}: {country}")
            seen.add(country)

            death_values = deaths[country]
            mmr_values = mmr[country]
            country_rows.append(
                {
                    "Region": region.display,
                    "Country": country,
                    "Death median total": as_output_number(death_values[0]),
                    "Death lower total": as_output_number(death_values[1]),
                    "Death upper total": as_output_number(death_values[2]),
                    "MMR median total": as_output_number(mmr_values[0]),
                    "MMR lower total": as_output_number(mmr_values[1]),
                    "MMR upper total": as_output_number(mmr_values[2]),
                    "Detectable deaths": death_values[1] > 0,
                    "Detectable MMR": mmr_values[1] > 0,
                }
            )

        death_total = add_triplets(list(deaths.values()))
        mmr_total = add_triplets(list(mmr.values()))
        global_death_triplets.append(death_total)
        global_mmr_triplets.append(mmr_total)
        all_countries.extend(combined_countries)

        table_rows_by_region.append(
            {
                "Region": region.display,
                "Countries included": len(combined_countries),
                "Cumulative excess maternal deaths": as_output_number(death_total[0]),
                "Deaths lower 95% UI": as_output_number(death_total[1]),
                "Deaths upper 95% UI": as_output_number(death_total[2]),
                "Formatted cumulative excess maternal deaths": fmt_triplet(*death_total),
                "Aggregate excess MMR": as_output_number(mmr_total[0]),
                "MMR lower 95% UI": as_output_number(mmr_total[1]),
                "MMR upper 95% UI": as_output_number(mmr_total[2]),
                "Formatted aggregate excess MMR": fmt_triplet(*mmr_total),
                "Countries with detectable excess deaths": sum(1 for values in deaths.values() if values[1] > 0),
                "Countries with detectable excess MMR": sum(1 for values in mmr.values() if values[1] > 0),
            }
        )

    global_death_total = add_triplets(global_death_triplets)
    global_mmr_total = add_triplets(global_mmr_triplets)
    global_row = {
        "Region": "Global",
        "Countries included": sum(row["Countries included"] for row in table_rows_by_region),
        "Cumulative excess maternal deaths": as_output_number(global_death_total[0]),
        "Deaths lower 95% UI": as_output_number(global_death_total[1]),
        "Deaths upper 95% UI": as_output_number(global_death_total[2]),
        "Formatted cumulative excess maternal deaths": fmt_triplet(*global_death_total),
        "Aggregate excess MMR": as_output_number(global_mmr_total[0]),
        "MMR lower 95% UI": as_output_number(global_mmr_total[1]),
        "MMR upper 95% UI": as_output_number(global_mmr_total[2]),
        "Formatted aggregate excess MMR": fmt_triplet(*global_mmr_total),
        "Countries with detectable excess deaths": sum(
            int(row["Countries with detectable excess deaths"]) for row in table_rows_by_region
        ),
        "Countries with detectable excess MMR": sum(
            int(row["Countries with detectable excess MMR"]) for row in table_rows_by_region
        ),
    }

    table_rows = [global_row] + table_rows_by_region

    duplicate_countries = sorted(country for country in set(all_countries) if all_countries.count(country) > 1)
    global_sum_confirmed = all(
        global_row[key] == sum(row[key] for row in table_rows_by_region)
        for key in [
            "Countries included",
            "Cumulative excess maternal deaths",
            "Deaths lower 95% UI",
            "Deaths upper 95% UI",
            "Aggregate excess MMR",
            "MMR lower 95% UI",
            "MMR upper 95% UI",
            "Countries with detectable excess deaths",
            "Countries with detectable excess MMR",
        ]
    )

    checks.append(("Total number of countries", global_row["Countries included"]))
    for row in table_rows_by_region:
        checks.append((f"Number of countries: {row['Region']}", row["Countries included"]))
    checks.append(("Death and MMR country sets match", "Yes" if all(set_matches) else "No"))
    checks.append(("Any missing countries", "; ".join(all_missing) if all_missing else "None"))
    checks.append(
        (
            "Any duplicated countries",
            "; ".join(all_duplicates or duplicate_countries) if (all_duplicates or duplicate_countries) else "None",
        )
    )
    checks.append(("Region-name corrections applied", "; ".join(corrections)))
    checks.append(("Global totals equal sum of regions", "Yes" if global_sum_confirmed else "No"))

    mismatches: list[str] = []
    for row in table_rows:
        expected = EXPECTED[row["Region"]]
        actual = (
            row["Countries included"],
            row["Cumulative excess maternal deaths"],
            row["Deaths lower 95% UI"],
            row["Deaths upper 95% UI"],
            row["Aggregate excess MMR"],
            row["MMR lower 95% UI"],
            row["MMR upper 95% UI"],
            row["Countries with detectable excess deaths"],
            row["Countries with detectable excess MMR"],
        )
        if tuple(actual) != expected:
            mismatches.append(f"{row['Region']}: expected {expected}, got {actual}")
    checks.append(("Matches expected audit numbers", "Yes" if not mismatches else "; ".join(mismatches)))

    wb = Workbook()
    ws_table = wb.active
    ws_table.title = "Table3"
    ws_country = wb.create_sheet("Country_Level_Check")
    ws_checks = wb.create_sheet("Checks")

    write_table3(ws_table, table_rows)
    write_country_level(ws_country, country_rows)
    write_checks(ws_checks, checks)

    for ws in [ws_table, ws_country, ws_checks]:
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30 if row_idx == 1 else 18
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor="D9EAF7")
            ws.cell(row=1, column=col_idx).alignment = Alignment(wrap_text=True, vertical="center")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)

    print(f"Wrote {OUTPUT_FILE}")
    print()
    print("| Region | Countries included | Cumulative excess maternal deaths | Aggregate excess MMR | Countries with detectable excess deaths | Countries with detectable excess MMR |")
    print("|---|---:|---:|---:|---:|---:|")
    for row in table_rows:
        print(
            f"| {row['Region']} | {row['Countries included']} | "
            f"{row['Formatted cumulative excess maternal deaths']} | "
            f"{row['Formatted aggregate excess MMR']} | "
            f"{row['Countries with detectable excess deaths']} | "
            f"{row['Countries with detectable excess MMR']} |"
        )


if __name__ == "__main__":
    main()
